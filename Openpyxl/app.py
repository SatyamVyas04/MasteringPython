import openpyxl as xl
# loading Workbook
wb = xl.load_workbook('transactions.xlsx')
# loading a singular sheet
sheet = wb["Sheet1"]
# assigning a variable the cell value
cell = sheet["a1"]  # or "cell = sheet.cell(1, 1)"

# printing rows and cols and also the assigned value
# print(sheet.max_row)
# print(sheet.max_column)
# print(cell.value)

# automatically changing the price (basic automation)
for row in range(2, sheet.max_row + 1): # starts from 2 so as to skip the title of the colm
    cell = sheet.cell(row, 3) # the price column
    cell_corrected = sheet.cell(row, 4) # corrected price column

    print("Price: ", end="")
    print(row, cell.value) # checking the current value in it

    corrected_price = cell.value*0.9 # whatever you want to change
    cell_corrected.value = corrected_price

    print("Corrected Price: ", end="")
    print(row, cell_corrected.value) # checking the current value in it

wb.save("transactions2.xlsx")